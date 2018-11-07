from configparser import ConfigParser
import datetime
import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter


__author__ = 'Dror Avrahami'
__version__ = '0.03'

# Change Log
# ----------
# 2018-10-18    0.01    Created
# 2018-10-21    0.02    Added comments and did some grooming. [Dror Av]
# 2018-10-21    0.03    Adjust width working properly and only once. [Dror Av]


class CreateExcel(object):
    """
    This class builds the Excel spreadsheet
    """

    def __init__(self):
        self.titles = ['IOC']
        self.MID_ALIGN = Alignment(wrap_text=True, vertical='center', horizontal='center')
        self.LEFT_ALIGN = Alignment(wrap_text=True, vertical='center', horizontal='left')
        self.GREY_FILL = PatternFill(start_color='f1f1f1', patternType='solid')
        self.WHITE_FONT = Font(bold=True, color='ffffff')
        self.BLACK_FILL = PatternFill(start_color='000000', patternType='solid')
        self.BORDER = Border(top=Side(color='a9a9a9', style='thin'), bottom=Side(color='a9a9a9', style='thin'),
                             left=Side(color='a9a9a9', style='thin'), right=Side(color='a9a9a9', style='thin'))

        self.table = {}
        self.workbook = Workbook()
        self.config = ConfigParser()
        self.logger = logging.getLogger('ExcelCreator')
        self._build_styles()

    def _build_styles(self):
        # Builds new styles for the spreadsheet

        title = NamedStyle("title")
        title.font = self.WHITE_FONT
        title.fill = self.BLACK_FILL
        title.alignment = self.MID_ALIGN

        self.workbook.add_named_style(title)

    def _adjust_width(self, sheet):
        # The Method Adjusts the width of the columns according to the longest string it finds (limit - 40)

        for column in sheet.iter_cols():
            max_length = 0
            col = column[0].column

            for cell in column:
                value = str(cell.value).split('\n')
                length = len(max(value, key=len))

                if length > max_length:
                    max_length = length

            if max_length > 40:
                max_length = 40

            sheet.column_dimensions[col].width = max_length

    def _apply_style_to_table(self, sheet, row):
        # The Method adds grey borders to rows and paints odd rows with grey fill.

        for column in range(1, len(self.titles)+1):
            cell = sheet['{}{}'.format(get_column_letter(column), row)]
            cell.border = self.BORDER

            if (row % 2) != 0:
                if cell.fill.start_color.index == '00000000':
                    cell.fill = self.GREY_FILL

        sheet.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(self.titles)), row)

        return 1

    def _build_titles(self, periods):
        # Builds the titles table according to the periods

        for days in periods:
            self.titles.extend(['{} Day\nResolutions'.format(days),
                                '{} Days\nMost common TLDs'.format(days),
                                '{} Days\nMost Common Domains'.format(days)])

    def _print_titles(self, titles, sheet, start_col=1):

        # ~~~ Print Titles ~~~

        for column, title in enumerate(titles, start_col):
            sheet.cell(column=column, row=1, value=title)
            self.table[title] = column

        # ~~~ Add Styles to Titles ~~~

        for col in range(1, len(titles)+1):
            c = sheet['{}1'.format(get_column_letter(col))]
            c.style = "title"

    def save_workbook(self, path, filename='tld_report'):
        """
        Saves the workbook as a local file and returns the absolute path
        :param path: Directory absolute location
        :param filename: Name for the saved file
        :return: absolute path of the file
        """

        dest_filename = '{filename} - {date:%Y-%m-%d %H%M%S}.xlsx'.format(filename=filename,
                                                                          date=datetime.datetime.today())
        try:
            self.logger.debug('Saving File {}'.format(path + dest_filename))
            self.workbook.save(filename=path + dest_filename)

        except IOError:

            self.logger.debug('Can\'t Find Folder \'{}\', Will try to Create it'.format(path))

            try:
                os.makedirs(path)
                self.logger.debug('Created Folder - \'{}\''.format(path))

                self.logger.debug('Saving File {}'.format(path + dest_filename))
                self.workbook.save(filename=path + dest_filename)

            except OSError as e:
                self.logger.error('Can\'t Create the Folder {}'.format(path))
                self.logger.error(e)
                exit()

            else:
                return path + dest_filename

    @staticmethod
    def _parse_data(data):
        # Parses the data in order to print it in the spreadsheet, also sorts it from max to min.

        top_lvl_domains = data['top_lvl_domains']

        second_lvl_domains = data['second_lvl_domains']

        top_lvl_domains = [(tld, top_lvl_domains[tld], top_lvl_domains[tld] / len(data['rdata']) * 100) for tld
                           in sorted(top_lvl_domains, key=top_lvl_domains.get, reverse=True)]

        second_lvl_domains = [(tld, second_lvl_domains[tld], second_lvl_domains[tld] / len(data['rdata']) * 100) for tld
                              in sorted(second_lvl_domains, key=second_lvl_domains.get, reverse=True)]

        return top_lvl_domains, second_lvl_domains

    @staticmethod
    def _build_top_list(data, max_tlds):
        # Formats the data as a string.

        return '\n'.join(['{} ({}, {:.2f}%)'.format(x[0], x[1], x[2]) for x in data[:max_tlds+1]])

    def run(self, iocs, periods, max_tlds):
        # The class main method

        self.logger.info('Creating an Excel File')

        sheet1 = self.workbook.active
        sheet1.title = 'TLD Breakdown'

        self._build_titles(periods)
        self._print_titles(self.titles, sheet1)

        # ~~~ Add Data to Table ~~~

        row = 2  # Starts printing the table data (row 1 is titles)

        for ioc in iocs:

            # ~~~ IOC ~~~
            sheet1.cell(column=1, row=row, value=ioc['ioc']).alignment = self.MID_ALIGN
            column = 2  # Column 1 is the IOC itself, this will restart the column index for each period.

            for period in periods:
                # Iterates over each period to add the relevant data

                # ~~~ Resolutions # ~~~

                sheet1.cell(column=column,
                            row=row,
                            value=len(ioc[period]['rdata'])).alignment = self.MID_ALIGN

                # ~~~ Most Common TLDs and Domains ~~~

                tlds, second_lvl_domains = self._parse_data(ioc[period])

                sheet1.cell(column=column+1,
                            row=row,
                            value=self._build_top_list(tlds, max_tlds)).alignment = self.LEFT_ALIGN

                sheet1.cell(column=column+2,
                            row=row,
                            value=self._build_top_list(second_lvl_domains, max_tlds)).alignment = self.LEFT_ALIGN

                column += 3  # Each period takes up 3 columns - Resolutions, TLDs, Domains

            # ~~~ Correcting column width and applying style ~~~

            self.logger.debug('Adjusting Column Width and Applying Table Style')
            self._apply_style_to_table(sheet1, row)
            row += 1  # Moves to the next row to handle the next IOC

        self._adjust_width(sheet1)


if __name__ == '__main__':
    pass
